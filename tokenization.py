import textblob

from textblob import TextBlob

import openpyxl

dataframe=openpyxl.load_workbook(dir)


dataframe1=dataframe.active

import nltk
nltk.download('punkt')


for row in range(1,dataframe1.max_row):
    for col in dataframe1.iter_cols(0,dataframe1.max_column):
        text=col[row].value
        blobObject=TextBlob(str(text))
        textWords=blobObject.words
        print(textWords)



