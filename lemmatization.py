
import pandas as pd

import openpyxl


dataframe = openpyxl.load_workbook(dir) 
#directory

dataframe1 = dataframe.active

from textblob import TextBlob

import nltk
nltk.download('punkt')
nltk.download('averaged_perceptron_tagger')
nltk.download('omw-1.4')
nltk.download('wordnet')
#necessary packages

def pos_tagger(text):
    sent = TextBlob(text)
    tag_dict = {"J": 'a', "N": 'n', "V": 'v', "R": 'r'}
    words_tags = [(w, tag_dict.get(pos[0], 'n')) for w, pos in sent.tags]   
    lemma_list = [wd.lemmatize(tag) for wd, tag in words_tags]
    return lemma_list
#lemmatization function


for row in range(1,dataframe1.max_row):
    for col in dataframe1.iter_cols(0,dataframe1.max_column):
        text=col[row].value
        lemma_list = pos_tagger(str(text))
        lemmatized_sentence = " ".join(lemma_list)
        print(lemmatized_sentence)


