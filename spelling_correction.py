import pandas as pd
import openpyxl

dataframe=openpyxl.load_workbook(dir)
#directory
dataframe1=dataframe.active

import textblob
from textblob import TextBlob

from spellchecker import SpellChecker

spell = SpellChecker()
        

for row in range(1,dataframe1.max_row):
    for col in dataframe1.iter_cols(0,dataframe1.max_column):
        text=col[row].value
        blobObject=TextBlob(str(text))
        textWords=blobObject.words
        misspelled = spell.unknown(textWords)
        for word in misspelled:
            for i in range(len(textWords)):
                if word==textWords[i]:
                    textWords[i]=spell.correction(word)
        print(textWords)


